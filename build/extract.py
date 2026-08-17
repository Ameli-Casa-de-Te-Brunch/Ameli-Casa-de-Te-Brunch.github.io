#!/usr/bin/env python3
"""xlsx (Ameli_Menu_Maestro, fuera del repo) -> data/menu.json (solo campos públicos).

Desde el maestro V3.1 el archivo tiene 4 hojas: 'Resumen y Configuración',
'Categorías', 'Productos' (nombre, descripción, precio, alérgenos — lo
esencial del menú) y 'Productos - Backoffice' (ingredientes y
personalización, referencia interna que el sitio nunca lee).
"""
import json
import re
import sys
from pathlib import Path
from urllib.parse import urlsplit

import openpyxl

import config_local

# Los 5 idiomas trabajados en el maestro.
LANGS = ["es", "en", "pt", "fr", "it"]

# Categorías cuyo segundo precio (cuando existe) se etiqueta "Vaso/Jarra" en
# vez de "Chico/Grande" — así lo pide el menú físico (batidos y jugos).
CATEGORIAS_VASO_JARRA = {"BYJ"}

# Estados de la columna "Estado de validación (alérgenos)" que habilitan
# publicar los datos de alérgenos de ESE producto puntual. Ver la nota legal
# en validate.py: mientras un producto diga otra cosa (por defecto,
# "Pendiente"), sus alérgenos nunca salen en el JSON público.
ESTADOS_ALERGENOS_VALIDADOS = {"Validado por cocina", "Validado por proveedor"}

# Claves cortas del alérgeno público <- columna del Excel.
_MAPA_ALERGENOS = [
    ("veg", "Vegetariano"), ("vgn", "Vegano"), ("tacc", "Sin TACC"), ("lac", "Sin lactosa"),
    ("glu", "Contiene gluten"), ("lech", "Contiene leche"), ("huev", "Contiene huevo"),
    ("soja", "Contiene soja"), ("mani", "Contiene maní"), ("fsec", "Contiene frutos secos"),
    ("ses", "Contiene sésamo"), ("pesc", "Contiene pescado"), ("mar", "Contiene mariscos"),
    ("alc", "Contiene alcohol"), ("caf", "Contiene cafeína"),
]

HERE = Path(__file__).resolve().parent
DEFAULT_OUT = HERE.parent / "data" / "menu.json"
OVERRIDES_MOMENTOS = HERE / "overrides_momentos.json"

FILA_CONFIG_INICIO = 16  # ver hoja "Resumen y Configuración": el bloque de
FILA_CONFIG_FIN = 29     # config empieza después del resumen automático.

# Columnas de la hoja "Productos" (1-indexado). Definidas una sola vez acá
# porque validate.py las necesita también para armar sus mensajes.
COL = {
    "id": 1, "cat": 2, "orden": 3,
    "nombre": {"es": 4, "en": 5, "pt": 6, "fr": 7, "it": 8},
    "desc": {"es": 9, "en": 10, "pt": 11, "fr": 12, "it": 13},
    "activo": 14, "destacado": 15, "recomendado": 16, "mas_vendido": 17, "nuevo": 18,
    "edicion_limitada": 19, "etiqueta_inicial": 20,
    "precio_chico": 21, "precio_grande": 22, "moneda": 23,
    "temperatura": 24, "formato": 25,
    "img": 26,
    "slug": {"es": 27, "en": 28, "pt": 29, "fr": 30, "it": 31},
    "alergenos_inicio": 32, "alergenos_fin": 46,
    "estado_alergenos": 47, "obs_alergenos": 48, "observaciones": 49,
}


def _sheet(wb, name):
    for candidate in wb.sheetnames:
        if candidate == name:
            return wb[candidate]
    raise KeyError(f"No se encontró la hoja {name!r}")


def load_categorias(wb):
    ws = _sheet(wb, "Categorías")
    cats = []
    r = 5
    while True:
        cod = ws.cell(row=r, column=1).value
        if cod is None:
            break
        visible = ws.cell(row=r, column=15).value
        cats.append({
            "cod": cod,
            "orden": ws.cell(row=r, column=2).value,
            "visible": visible == "Sí",
            "nom": {
                "es": ws.cell(row=r, column=9).value,
                "en": ws.cell(row=r, column=10).value,
                "pt": ws.cell(row=r, column=11).value,
                "fr": ws.cell(row=r, column=12).value,
                "it": ws.cell(row=r, column=13).value,
            },
        })
        r += 1
    cats.sort(key=lambda c: c["orden"])
    return cats


def _formatear_precio(chico, grande, cat_cod):
    """Un solo precio -> "$ X". Dos precios -> "Chico $ X · Grande $ Y"
    (o "Vaso/Jarra" para batidos y jugos), según lo que haya cargado."""
    if chico in (None, "") and grande in (None, ""):
        return None
    et1, et2 = ("Vaso", "Jarra") if cat_cod in CATEGORIAS_VASO_JARRA else ("Chico", "Grande")
    fmt = lambda v: f"$ {v:,.0f}".replace(",", ".")
    if grande in (None, ""):
        return fmt(chico)
    if chico in (None, ""):
        return f"{et2} {fmt(grande)}"
    return f"{et1} {fmt(chico)} · {et2} {fmt(grande)}"


def _equivalente(precio_ars, tasa, simbolo):
    """Convierte el precio de referencia (el chico/único, nunca el grande —
    para no saturar la tarjeta con cuatro números) a otra moneda, usando la
    tasa manual del Excel. Ninguna llamada externa: es la conversión fija
    del día que se publicó, no una cotización en vivo (ver README, sección
    de cabeceras de seguridad, para la razón: no queríamos abrir connect-src
    a un tercero por esto)."""
    if precio_ars in (None, "") or tasa in (None, "") or tasa == 0:
        return None
    valor = precio_ars / tasa
    return f"≈ {simbolo} {valor:,.0f}".replace(",", ".")


def load_opciones_leche(wb):
    """ID -> lista de opciones de leche disponibles ("veg", "lac"), leídas de
    'Productos - Backoffice' (columnas 'Leche vegetal' / 'Leche sin lactosa').
    Esto reemplaza a los viejos productos ADI001/ADI002 ("Leche vegetal",
    "Leche sin lactosa"): no son productos que se pidan solos, son un
    agregado para las bebidas que ya llevan leche — se muestra en el
    detalle del producto correspondiente, no como su propia fila de menú."""
    try:
        ws = wb["Productos - Backoffice"]
    except KeyError:
        return {}
    COL_LECHE_VEGETAL, COL_LECHE_SIN_LACTOSA = 18, 19
    out = {}
    r = 5
    while True:
        idv = ws.cell(row=r, column=1).value
        if idv is None:
            break
        ops = []
        if ws.cell(row=r, column=COL_LECHE_VEGETAL).value == "Sí":
            ops.append("veg")
        if ws.cell(row=r, column=COL_LECHE_SIN_LACTOSA).value == "Sí":
            ops.append("lac")
        if ops:
            out[idv] = ops
        r += 1
    return out


def load_productos(wb):
    """ID -> todos los datos de producto que usa el pipeline, en un solo lugar."""
    ws = _sheet(wb, "Productos")
    opciones_leche = load_opciones_leche(wb)
    out = {}
    r = 5
    while True:
        idv = ws.cell(row=r, column=COL["id"]).value
        if idv is None:
            break
        cat_cod = ws.cell(row=r, column=COL["cat"]).value
        chico = ws.cell(row=r, column=COL["precio_chico"]).value
        grande = ws.cell(row=r, column=COL["precio_grande"]).value
        estado_alergenos = ws.cell(row=r, column=COL["estado_alergenos"]).value

        alergenos = None
        if estado_alergenos in ESTADOS_ALERGENOS_VALIDADOS:
            c = COL["alergenos_inicio"]
            alergenos = {}
            for clave, _nombre in _MAPA_ALERGENOS:
                alergenos[clave] = ws.cell(row=r, column=c).value == "Sí"
                c += 1

        out[idv] = {
            "cat": cat_cod,
            "orden": ws.cell(row=r, column=COL["orden"]).value,
            "n": {lang: ws.cell(row=r, column=c).value for lang, c in
                  ((l, COL["nombre"][l]) for l in LANGS)},
            "d": {lang: ws.cell(row=r, column=c).value for lang, c in
                  ((l, COL["desc"][l]) for l in LANGS)},
            "activo": ws.cell(row=r, column=COL["activo"]).value == "Sí",
            "destacado": ws.cell(row=r, column=COL["destacado"]).value == "Sí",
            "recomendado": ws.cell(row=r, column=COL["recomendado"]).value == "Sí",
            "mas_vendido": ws.cell(row=r, column=COL["mas_vendido"]).value == "Sí",
            "nuevo": ws.cell(row=r, column=COL["nuevo"]).value == "Sí",
            "precio": _formatear_precio(chico, grande, cat_cod),
            "precio_chico_ars": chico if chico not in (None, "") else grande,
            "temperatura": ws.cell(row=r, column=COL["temperatura"]).value or "",
            "formato": ws.cell(row=r, column=COL["formato"]).value or "",
            "img": ws.cell(row=r, column=COL["img"]).value or None,
            "tag": ws.cell(row=r, column=COL["etiqueta_inicial"]).value or None,
            "alerg": alergenos,
            "estado_alergenos": estado_alergenos,
            "leche": opciones_leche.get(idv, []),
            "_fila": r,
        }
        r += 1
    return out


def _normalizar_whatsapp(valor):
    """La celda puede venir como número (Excel la interpreta así si son solo dígitos)
    o como texto con espacios/guiones. Siempre devolvemos solo dígitos, o None."""
    if valor in (None, ""):
        return None
    if isinstance(valor, float) and valor.is_integer():
        valor = int(valor)
    solo_digitos = re.sub(r"\D", "", str(valor))
    return solo_digitos or None


def _es_placeholder(valor):
    """Detecta placeholders obvios tipo 'XXXXXXXX' o 'ejemplo' que no deberían publicarse."""
    if valor in (None, ""):
        return False
    texto = str(valor).strip().lower()
    return "xxx" in texto or texto in ("ejemplo", "pendiente", "completar", "tbd", "n/a")


def _url_https_valida(valor, dominios_permitidos=None):
    """Antes de publicar un link que viene del Excel como texto libre (no un
    handle ni un teléfono que ya sanitizamos con regex), lo validamos: solo
    https, y si se pasa una lista de dominios, el host tiene que contener
    alguno de ellos. Esto es una segunda capa además del escapado HTML en
    render.py — no confiamos en que la celda del Excel siempre tenga lo que
    se espera (podría pegarse mal, quedar a medio escribir, etc.), y un
    esquema no-https (`javascript:`, `data:`, ...) nunca debería llegar a
    un href publicado, más allá de que el CSP también lo bloquee."""
    if valor in (None, ""):
        return False
    texto = str(valor).strip()
    try:
        partes = urlsplit(texto)
    except ValueError:
        return False
    if partes.scheme != "https" or not partes.netloc:
        return False
    if dominios_permitidos and not any(d in partes.netloc.lower() for d in dominios_permitidos):
        return False
    return True


def _handle_instagram_sano(valor):
    """Los handles de Instagram son solo letras, números, punto y guion bajo.
    Cualquier otra cosa (comillas, ángulos, espacios) no es un handle válido
    y no debería terminar pegada sin escapar dentro de un href — se
    descarta en vez de intentar 'arreglarla'."""
    if valor in (None, ""):
        return None
    texto = str(valor).strip().lstrip("@")
    if not texto or not re.fullmatch(r"[A-Za-z0-9._]{1,30}", texto):
        return None
    return texto


def load_config(wb):
    ws = _sheet(wb, "Resumen y Configuración")
    params = {}
    for r in range(FILA_CONFIG_INICIO, FILA_CONFIG_FIN + 1):
        nombre = ws.cell(row=r, column=1).value
        if nombre is None:
            continue
        params[str(nombre).strip()] = ws.cell(row=r, column=2).value

    whatsapp = _normalizar_whatsapp(params.get("WhatsApp de pedidos"))
    instagram = params.get("Instagram")
    direccion = params.get("Dirección")
    url_base = params.get("URL base del menú")
    tripadvisor = params.get("TripAdvisor")
    google_resenas = params.get("Google (reseñas)")

    if _es_placeholder(whatsapp) or _es_placeholder(params.get("WhatsApp de pedidos")):
        whatsapp = None
    if _es_placeholder(instagram):
        instagram = None
    if _es_placeholder(direccion):
        direccion = None
    if _es_placeholder(url_base):
        url_base = None
    if _es_placeholder(tripadvisor):
        tripadvisor = None
    if _es_placeholder(google_resenas):
        google_resenas = None

    # Segunda capa de validación (además del escapado HTML en render.py):
    # un handle de Instagram con caracteres raros, o un link que no sea
    # https con el dominio esperado, no se publica — mejor ausente que
    # roto o, peor, colado sin escapar en un atributo href.
    instagram = _handle_instagram_sano(instagram)
    if not _url_https_valida(url_base):
        url_base = None
    if not _url_https_valida(tripadvisor, ("tripadvisor.",)):
        tripadvisor = None
    if not _url_https_valida(google_resenas, ("google.com", "g.page")):
        google_resenas = None

    tasa_usd = params.get("Tipo de cambio ARS/USD")
    tasa_eur = params.get("Tipo de cambio ARS/EUR")
    tasa_brl = params.get("Real")

    return {
        "moneda": params.get("Moneda local") or "ARS",
        "whatsapp": whatsapp,
        "instagram": instagram,
        "direccion": direccion,
        "url_base": url_base,
        "tripadvisor": tripadvisor,
        "google_resenas": google_resenas,
        "tasa_usd": tasa_usd if isinstance(tasa_usd, (int, float)) else None,
        "tasa_eur": tasa_eur if isinstance(tasa_eur, (int, float)) else None,
        "tasa_brl": tasa_brl if isinstance(tasa_brl, (int, float)) else None,
    }


def moments_for(prod, overrides):
    temp = prod["temperatura"]
    formato = prod["formato"]
    cat_cod = prod["cat"]
    m = []
    if "Caliente" in temp:
        m.append("calentito")
    if "Fría" in temp:
        m.append("fresco")
    if "jarra" in formato or cat_cod == "TYT":
        m.append("compartir")
    if "Unidad" in formato or cat_cod == "STC":
        m.append("llevar")
    return m


def badges_for(prod):
    b = []
    if prod["destacado"]:
        b.append("fav")
    elif prod["recomendado"]:
        b.append("reco")
    elif prod["mas_vendido"]:
        b.append("pedido")
    elif prod["nuevo"]:
        b.append("nuevo")
    if prod["cat"] == "STC":
        b.append("sintacc")
    return b


def extract(xlsx_path: Path) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    cats = load_categorias(wb)
    productos = load_productos(wb)
    config = load_config(wb)
    overrides = json.loads(OVERRIDES_MOMENTOS.read_text(encoding="utf-8"))["extra"]

    prods = []
    precios = {}
    meta = {}
    # metadata de TODOS los productos (activos o no), para que el validador
    # pueda señalar filas concretas incluso en productos inactivos
    for prod_id, prod in productos.items():
        meta[prod_id] = {
            "fila": prod["_fila"],
            "cat": prod["cat"],
            "activo": prod["activo"],
            "destacado": prod["destacado"],
            "nombre_es": prod["n"]["es"] or prod_id,
        }

    for prod_id, prod in productos.items():
        if not prod["activo"]:
            continue
        m = moments_for(prod, overrides)
        for extra in overrides.get(prod_id, []):
            if extra not in m:
                m.append(extra)
        if prod["precio"]:
            entrada = {"ars": prod["precio"]}
            usd = _equivalente(prod["precio_chico_ars"], config["tasa_usd"], "USD")
            eur = _equivalente(prod["precio_chico_ars"], config["tasa_eur"], "EUR")
            brl = _equivalente(prod["precio_chico_ars"], config["tasa_brl"], "R$")
            if usd:
                entrada["usd"] = usd
            if eur:
                entrada["eur"] = eur
            if brl:
                entrada["brl"] = brl
            precios[prod_id] = entrada
        item = {
            "id": prod_id,
            "cat": prod["cat"],
            "orden": prod["orden"],
            "dest": prod["destacado"],
            "n": prod["n"],
            "d": prod["d"],
            "m": m,
            "b": badges_for(prod),
            "img": prod["img"],
        }
        if prod["alerg"] is not None:
            item["alerg"] = prod["alerg"]
        if prod["tag"]:
            item["tag"] = prod["tag"]
        if prod["leche"]:
            item["leche"] = prod["leche"]
        prods.append(item)
    prods.sort(key=lambda p: (next(c["orden"] for c in cats if c["cod"] == p["cat"]), p["orden"]))

    cats_out = [{"cod": c["cod"], "orden": c["orden"], "nom": c["nom"]} for c in cats if c["visible"]]

    return {
        "cats": cats_out,
        "prods": prods,
        "precios": precios,
        "config": config,
        "_meta": meta,
    }


# Campos que SÍ salen al sitio público. "alerg" solo aparece en un producto
# si su fila individual está validada (ver load_productos) — el resto
# (costos, ingredientes, personalización, notas operativas, fila/columna
# de origen) se queda afuera de lo que se versiona y se publica.
_CAMPOS_PROD_PUBLICOS = ("id", "cat", "orden", "dest", "n", "d", "m", "b", "img", "alerg", "tag", "leche")
_CAMPOS_CONFIG_PUBLICOS = ("moneda", "whatsapp", "instagram", "direccion", "url_base", "tripadvisor", "google_resenas")


def datos_publicos(data: dict) -> dict:
    """Proyección de extract() con solo lo que un visitante del menú necesita ver.
    Esto es lo que se escribe a disco y se versiona — nunca el dict completo
    (que trae _meta: filas de origen, flags internos, etc. útiles solo para
    que validate.py arme sus mensajes en el mismo proceso)."""
    return {
        "cats": data["cats"],
        "prods": [{k: p[k] for k in _CAMPOS_PROD_PUBLICOS if k in p} for p in data["prods"]],
        "precios": data["precios"],
        "config": {k: data["config"].get(k) for k in _CAMPOS_CONFIG_PUBLICOS},
    }


def main():
    xlsx_arg = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = config_local.resolver_ruta_xlsx(xlsx_arg)
    if not xlsx_path.exists():
        print(config_local.mensaje_no_encontrado(xlsx_path))
        sys.exit(1)
    out_path = Path(sys.argv[2]) if len(sys.argv) > 2 else DEFAULT_OUT
    data = extract(xlsx_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(datos_publicos(data), ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"OK: {len(data['prods'])} productos activos, {len(data['cats'])} categorías -> {out_path}")


if __name__ == "__main__":
    main()
