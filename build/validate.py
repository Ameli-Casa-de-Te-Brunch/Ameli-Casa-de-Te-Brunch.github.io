#!/usr/bin/env python3
"""Valida el maestro antes de publicar.

Habla en castellano llano: qué pasa, en qué hoja, en qué fila. El dueño del
negocio no es programador, así que cada mensaje dice exactamente qué corregir
y dónde.

Errores -> bloquean el build (build.py corta y no genera dist/index.html).
Warnings -> se listan pero no bloquean (ej. todavía no hay precios cargados,
o un producto activo todavía no tiene sus alérgenos validados).

Desde el maestro V3.1 el archivo tiene 4 hojas: 'Resumen y Configuración',
'Categorías', 'Productos' (nombre, descripción, precio, alérgenos) y
'Productos - Backoffice' (ingredientes y personalización, referencia interna).
"""
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

import config_local
import extract

HERE = Path(__file__).resolve().parent

# Solo 3 idiomas exigidos por el validador — ver la misma nota en extract.py.
LANGS = ["es", "en", "pt"]
IDIOMA_NOMBRE = {"es": "español", "en": "inglés", "pt": "portugués"}

ID_FORMATO = re.compile(r"^[A-Z]{3}\d{3}$")

CAMPOS_CONTACTO = {"WhatsApp de pedidos", "Instagram", "Dirección", "URL base del menú"}


def _col(nombre_campo, lang=None):
    c = extract.COL[nombre_campo]
    if lang:
        c = c[lang]
    return get_column_letter(c)


def _es_placeholder(valor):
    if valor in (None, ""):
        return False
    texto = str(valor).strip().lower()
    return "xxx" in texto or texto in ("ejemplo", "pendiente", "completar", "tbd", "n/a")


def _leer_ids_productos(wb):
    """Todas las filas de Productos, incluso con IDs repetidos o mal formados
    (a diferencia de extract.py, que ya los usa como clave de diccionario y pierde duplicados)."""
    ws = wb["Productos"]
    filas = []
    r = 5
    while True:
        idv = ws.cell(row=r, column=1).value
        if idv is None:
            break
        filas.append((r, idv))
        r += 1
    return filas


def _leer_slugs(wb):
    ws = wb["Productos"]
    filas = []
    r = 5
    while True:
        idv = ws.cell(row=r, column=1).value
        if idv is None:
            break
        filas.append((r, idv, ws.cell(row=r, column=extract.COL["slug"]["es"]).value,
                       ws.cell(row=r, column=extract.COL["slug"]["en"]).value))
        r += 1
    return filas


def _leer_config_crudo(wb):
    ws = wb["Resumen y Configuración"]
    valores = {}
    for r in range(extract.FILA_CONFIG_INICIO, extract.FILA_CONFIG_FIN + 1):
        nombre = ws.cell(row=r, column=1).value
        if nombre is None:
            continue
        valores[nombre] = (r, ws.cell(row=r, column=2).value)
    return valores


def validate(data: dict, xlsx_path: Path):
    errors, warnings = [], []
    if not xlsx_path.exists():
        errors.append(f"No encuentro el archivo {xlsx_path} — revisá la ruta.")
        return errors, warnings

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    meta = data.get("_meta", {})
    cat_codes = {c["cod"] for c in data["cats"]}

    # --- IDs: formato y duplicados (sobre TODAS las filas de Productos, activas o no) ---
    filas_productos = _leer_ids_productos(wb)
    vistos = {}
    for fila, idv in filas_productos:
        idv_txt = str(idv).strip()
        if not ID_FORMATO.match(idv_txt):
            errors.append(
                f"Fila {fila} de Productos: el ID '{idv_txt}' no tiene el formato "
                f"esperado (3 letras + 3 números, ej. TYT004). Corregilo en la columna A."
            )
        if idv_txt in vistos:
            errors.append(
                f"El ID '{idv_txt}' está repetido en Productos: filas {vistos[idv_txt]} "
                f"y {fila}. Cada producto necesita un ID único — cambiá uno de los dos."
            )
        else:
            vistos[idv_txt] = fila

    # --- por producto activo: traducciones, categoría, precio, alérgenos ---
    for p in data["prods"]:
        pid = p["id"]
        info = meta.get(pid, {})
        nombre = info.get("nombre_es") or pid
        fila = info.get("fila")

        if p["cat"] not in cat_codes:
            errors.append(
                f"Fila {fila} ({pid} · {nombre}): la categoría '{p['cat']}' (columna "
                f"{_col('cat')}) no existe en la hoja Categorías, o no está marcada "
                f"como visible ahí."
            )

        for lang in LANGS:
            if not p["n"].get(lang):
                errors.append(
                    f"Fila {fila} ({pid} · {nombre}): falta el nombre en {IDIOMA_NOMBRE[lang]}. "
                    f"Cargalo en la hoja Productos, columna {_col('nombre', lang)}."
                )
            if not p["d"].get(lang):
                errors.append(
                    f"Fila {fila} ({pid} · {nombre}): falta la descripción en {IDIOMA_NOMBRE[lang]}. "
                    f"Cargala en la hoja Productos, columna {_col('desc', lang)}."
                )

        if pid not in data["precios"]:
            warnings.append(
                f"Fila {fila} ({pid} · {nombre}): está activo pero sin precio cargado en "
                f"la hoja Productos, columna {_col('precio_chico')} (o su columna de precio "
                f"grande) — se va a publicar sin precio visible."
            )

        if "alerg" not in p:
            warnings.append(
                f"Fila {fila} ({pid} · {nombre}): sus alérgenos todavía no están validados "
                f"(columna {_col('estado_alergenos')} de Productos tiene que decir 'Validado "
                f"por cocina' o 'Validado por proveedor') — se va a publicar sin esa "
                f"información hasta entonces."
            )

    # --- slugs duplicados ---
    filas_slugs = _leer_slugs(wb)
    for idx_col, label in ((2, "ES"), (3, "EN")):
        vistos_slug = {}
        for fila, idv, slug_es, slug_en in filas_slugs:
            slug = slug_es if idx_col == 2 else slug_en
            if not slug:
                continue
            if slug in vistos_slug:
                fila_prev, id_prev = vistos_slug[slug]
                errors.append(
                    f"El slug {label} '{slug}' se repite en Productos: fila {fila_prev} "
                    f"({id_prev}) y fila {fila} ({idv}). Cada producto necesita un slug único."
                )
            else:
                vistos_slug[slug] = (fila, idv)

    # --- datos de contacto: placeholders detectados directamente en el bloque de config ---
    config_crudo = _leer_config_crudo(wb)
    for campo, (fila, valor) in config_crudo.items():
        if campo not in CAMPOS_CONTACTO:
            continue
        if valor in (None, ""):
            warnings.append(
                f"Hoja Resumen y Configuración, fila {fila} ('{campo}'): está vacío. El elemento "
                f"correspondiente no se va a mostrar en el sitio hasta que lo completes."
            )
        elif _es_placeholder(valor):
            warnings.append(
                f"Hoja Resumen y Configuración, fila {fila} ('{campo}'): el valor '{valor}' parece un "
                f"dato de ejemplo, no uno real — no se va a publicar hasta que lo reemplaces."
            )

    return errors, warnings


def main():
    xlsx_arg = sys.argv[1] if len(sys.argv) > 1 else None
    xlsx_path = config_local.resolver_ruta_xlsx(xlsx_arg)
    if not xlsx_path.exists():
        print(config_local.mensaje_no_encontrado(xlsx_path))
        sys.exit(1)
    data = extract.extract(xlsx_path)  # incluye _meta, necesario para las filas de los mensajes
    errors, warnings = validate(data, xlsx_path)

    for w in warnings:
        print(f"[AVISO]  {w}")
    for e in errors:
        print(f"[ERROR] {e}")

    print(f"\n{len(errors)} error(es), {len(warnings)} aviso(s).")
    if errors:
        sys.exit(1)


if __name__ == "__main__":
    main()
